import pandas as pd
import numpy as np

import math

from openpyxl import load_workbook
workbook = load_workbook(filename="C:\\Yapay_Zeka_Algoritma_Uygulamalari\\naive_bayes_metabric-main\\datauji.xlsx")

sheet = workbook.active

from sklearn.model_selection import train_test_split

# Get neccesary columns
datarule=pd.read_excel('rulesNaive.xls', sheet_name = 'Sheet1')

arr_datarule = datarule.to_numpy()
#Sayısal sütunlar için ortalama ve standart sapmaları elde etmek için iki fonksiyon tanımlandı
def calculate_numeric(x, mean, standar):
    return 1 / (math.sqrt( math.pi * 2 * standar )  ) * math.exp(-( pow((x-mean), 2)/ (2 * pow(standar, 2)) ))


 # Her bir sayısal öznitelik için uygun ortalama değerleri al
def getmean_numeric(rules, header):
    result = {
      "living": 0.0,
      "died": 0.0
    }
    #header öznitelik
    if header == "tumor_size":
        result["living"]= rules[0][2]#ilk satır ikinci sütun: yaşayan ("living") hastaların tumor boyutunun ortalama değeridir.
        result["died"]= rules[0][3]#ilk satır 3.sütun : ölen ("died") hastaların tumor boyutunun ortalama değeridir.
    elif header == "tumor_stage":
        result["living"]= rules[2][2]#yaşayan ("living") hastaların tumor evresinin ortalama değeridir.
        result["died"]= rules[2][3]#ölen ("died") hastaların tumor evresinin ortalama değeridir.

    elif header == "mutation_count":
        result["living"]= rules[18][2]#Yaşayan ("died") hastaların mutasyon sayısının ortalama değeridir.
        result["died"]= rules[18][3]#ölen ("died") hastaların mutasyon sayısının ortalama değeridir.
    elif header == "neoplasm_histologic_grade":
        result["living"]= rules[34][2]
        result["died"]= rules[34][3]
        
    elif header == "age_at_diagnosis":
        result["living"]= rules[65][2]#yaşayan ("living") hastaların tanı anındaki yaşının ortalama değeridir.
        result["died"]= rules[65][3]#ölen ("living") hastaların tanı anındaki yaşının ortalama değeridir.
        
    elif header == "cohort":#belirli bir zaman diliminde tanı konmuş olan hastaları içerebilir.
        result["living"]= rules[78][2]
        result["died"]= rules[78][3]
        
    elif header == "lymph_nodes_examined_positive":
        result["living"]= rules[106][2]
        result["died"]= rules[106][3]
    elif header == "nottingham_prognostic_index":
        result["living"]= rules[108][2]
        result["died"]= rules[108][3]
    return result

 # Her bir sayısal öznitelik için uygun ortalama standart sapmasını al 
def getstandar_numeric(rules, header):
    result = {
      "living": 0.0,
      "died": 0.0
    }
    if header == "tumor_size":
        result["living"]= rules[1][2]
        result["died"]= rules[1][3]
    elif header == "tumor_stage":
        result["living"]= rules[3][2]
        result["died"]= rules[3][3]
    elif header == "mutation_count":
        result["living"]= rules[19][2]
        result["died"]= rules[19][3]
    elif header == "neoplasm_histologic_grade":
        result["living"]= rules[35][2]
        result["died"]= rules[35][3]
    elif header == "age_at_diagnosis":
        result["living"]= rules[66][2]
        result["died"]= rules[66][3]
    elif header == "cohort":
        result["living"]= rules[79][2]
        result["died"]= rules[79][3]
    elif header == "lymph_nodes_examined_positive":
        result["living"]= rules[107][2]
        result["died"]= rules[107][3]
    elif header == "nottingham_prognostic_index":
        result["living"]= rules[109][2]
        result["died"]= rules[109][3]
    return result

def getProbs_nominal(rules, header, value):
    # Boş bir sözlük oluşturulur ve başlangıç değerleri 0.0 olarak atanır.
    result = {
      "living": 0.0,
      "died": 0.0
    }
    # Koşul ifadesi, value değişkeninin string formatına dönüştürülüp, 
    # header adıyla birleştirilerek condition değişkenine atanır.
    condition = "value={}".format(value)
    
    # Eğer header özniteliği sayısal öznitelikler listesinde değilse:
    if header not in {"tumor_size", "tumor_stage", "mutation_count", "neoplasm_histologic_grade", "age_at_diagnosis", "cohort",
    "lymph_nodes_examined_positive", "nottingham_prognostic_index"}:
        
        # rules listesindeki her bir kural için döngü oluşturulur.
        for x in rules:
            # Eğer kuralın ilk öğesi (x[0]) header ile aynı ve 
            # ikinci öğesi (x[1]) condition ile aynıysa:
            if x[0] == header and x[1] == condition:
                # result sözlüğündeki living anahtarına kuralın üçüncü öğesi (x[2]) atanır.
                result["living"] = x[2]
                
                # result sözlüğündeki died anahtarına kuralın dördüncü öğesi (x[3]) atanır.
                result["died"] = x[3]
                
    # Sonuç olarak, result sözlüğü döndürülür.
    return result

tophead = ()
for value in sheet.iter_rows(max_row=1, #Excel sayfasının ilk satırındaki verileri (max_row=1) alınır.
                              min_col=1,
                              max_col=29,
                              values_only=True):
   tophead = value

row = 1
disscorrect = 0

# Excel sayfasındaki veriler satır bazında işlenir.
# Her bir satır, veriler sadece değer olarak alınarak iterasyon yapılır.
for value in sheet.iter_rows(min_row=2,#kinci satırından başlayarak (min_row=2), 29 sütuna kadar olan veriler alınır 
                              min_col=1,
                              max_col=29,
                              values_only=True):
   row = row + 1
   
   # Yaşayan ve ölen olasılıklar başlangıçta 1 olarak tanımlanır.
   probs_living = float(1)
   probs_died = float(1)
   for idx in range(len(value)):
       
        # Eğer öznitelik sayısal ise:
       if idx in {0, 7, 10, 18, 19, 20, 25, 26}:
           
           # Ortalama ve standart sapma değerleri hesaplanır.
           mean = getmean_numeric(arr_datarule, tophead[idx])
           standar = getstandar_numeric(arr_datarule, tophead[idx])
           
           # Eğer öznitelik değeri boşsa:
           if value[idx] is None:
               # Ortalama ve standart sapma kullanılarak olasılıklar hesaplanır.
               probs_living = probs_living * calculate_numeric(mean['living'], mean['living'], standar['living'])
               probs_died = probs_died * calculate_numeric(mean['died'], mean['died'], standar['died'])
           else:
               
               # Değer kullanılarak olasılıklar hesaplanır.
               valnumeric = value[idx]
               probs_living = probs_living * calculate_numeric(valnumeric, mean['living'], standar['living'])
               probs_died = probs_died * calculate_numeric(valnumeric, mean['died'], standar['died'])
        # Eğer öznitelik kategorik ise:
       elif idx in {1,2,3,4,5,6,8,9,11,12,13,14,15,16,17,21,22,23,24}:
           
           # Öznitelik değeri belirli bir formata dönüştürülür.
           if value[idx] is None:
               valnominal = "UNDEF"
           elif value[idx] == 0:
               valnominal = "false"
           elif value[idx] == 1:
               valnominal = "true"
           else:
               valnominal = value[idx]
               
            # Belirli bir kategorik öznitelik için yaşayan ve ölen olasılıklar hesaplanır.
           probablilitas_evidence = getProbs_nominal(arr_datarule, tophead[idx], valnominal)
           probs_living = probs_living * probablilitas_evidence["living"]
           probs_died = probs_died * probablilitas_evidence["died"]
           
    # Sınıf tahmini yapılır (Yaşayan mı, hastalıktan mı ölmüş?)
   sheet['AC{}'.format(row)] = probs_living
   sheet['AD{}'.format(row)] = probs_died
   classpred = None
   if probs_living > probs_died:
       classpred = "Living"
   else:
       classpred = "Died of Disease"
   sheet['AE{}'.format(row)] = classpred


workbook.save("datauji.xlsx")
